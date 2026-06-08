"""
core/roi_calculator.py — ROI 산정 엔진
======================================
엑셀 시트의 산정식을 Python 함수로 변환.

데이터 소스:
    - 07_조달청_단가DB.xlsx (442개 자재, 14개 GR 카테고리)
    - 08_조달청_간접공사비_2026.xlsx (공사규모×기간 매트릭스)
    - 01_GR_가이드라인.pdf (보조율 50%/70%)
    - 04_녹색건축법.pdf (§15 용적률 115% 완화)
    - 05_지방세특례제한법.pdf (§47의2 취득세 15~20% 감면)

검증 목표 (KEPCO 김천 도담어린이집):
    Max Cost 2.92억 / 자부담 1.46억 / 자산화 ROI 16.9%
"""

import os
from typing import Optional

import pandas as pd
from dotenv import load_dotenv

load_dotenv()


# ====================================================================
# 상수 정의 (자료 04·05 + 01 가이드라인 기반)
# ====================================================================

# 녹색건축법 §15 + 시행령 - 용적률 완화 (ZEB 등급별)
FAR_BONUS_BY_GRADE = {
    1: 0.15,   # 1등급: 최대 15% (법령 상한)
    2: 0.14,
    3: 0.13,
    4: 0.12,
    5: 0.11,   # 5등급: 11%
}

# 지방세특례 §47의2 - ZEB 인증 취득세 감면율
ZEB_TAX_RELIEF_RATE = {
    1: 0.20,
    2: 0.20,
    3: 0.18,
    4: 0.18,
    5: 0.15,
}

# 업무시설 취득세율
ACQUISITION_TAX_RATE = 0.022

# GR 가이드라인 - 보조율
SUBSIDY_RATE = {
    "서울_중앙_공공": 0.50,
    "그외_지자체": 0.70,
}

# GR 가이드라인 - 단위면적당 지원한도 (만원/3.3㎡)
SUBSIDY_LIMIT_PER_PYEONG = {
    "서울_중앙_공공_일반": 200,
    "그외_지자체_일반": 280,
    "서울_중앙_공공_소규모": 210,
    "그외_지자체_소규모": 294,
    "서울_중앙_공공_시그니처": 400,
    "그외_지자체_시그니처": 560,
}

# GR 가이드라인 - 사업당 국비 지원 상한
SUBSIDY_PROJECT_CAP = {
    "서울_중앙_공공": 5_000_000_000,
    "그외_지자체": 7_000_000_000,
}

# --------------------------------------------------------------------
# 시공·마감·가설 보정 계수 (CONSTRUCTION_FACTOR)
# --------------------------------------------------------------------
# 조달청 단가DB는 "자재 단가"만 포함. 실제 공사비는 자재 외에 다음을 포함:
#   - 시공 노무비 (직노)
#   - 가설공사 (비계, 안전망 등)
#   - 부속 자재 (단열재 고정장치, 방수재, 코킹 등)
#   - 마감재 (외단열 시 석고보드·도장 등)
#
# 본 상수는 졸업설계 단계의 보수적 추정값입니다.
# 실시설계 단계에선 견적사·시공사 의견을 반영하여 갱신해야 합니다.
# 출처: 한국건설기술연구원 표준품셈 평균 비율 추정
# --------------------------------------------------------------------
CONSTRUCTION_FACTOR = {
    "외벽_외단열": 3.0,       # 단열재 × 3 = 종합 공사비 (≈ 101,400원/㎡)
    "지붕_단열": 2.5,         # 지붕은 마감이 단순
    "바닥_단열": 2.5,
    "창호": 11.0,             # 유리 단가 × 11 (프레임+시공이 자재의 10배)
    "단열문": 1.5,            # 금속문은 완성품이라 시공비 비중 낮음
}


# ====================================================================
# Phase A - 데이터 로더
# ====================================================================

def load_price_db(xlsx_path: Optional[str] = None) -> pd.DataFrame:
    """07 조달청 단가DB -> pandas DataFrame."""
    if xlsx_path is None:
        xlsx_path = os.getenv(
            "PRICE_DB_PATH",
            "./data/policy_docs/07_조달청_단가DB.xlsx",
        )

    df = pd.read_excel(xlsx_path, sheet_name="3_단가DB_조달청", header=4)

    df = df.rename(columns={
        "GR 카테고리": "GR_카테고리",
        "규격 요약": "규격_요약",
        "규격 상세": "규격_상세",
        "가격(원)": "가격_원",
        "두께(mm)": "두께_mm",
        "열전도율(W/mK)": "열전도율",
    })

    df = df[df["GR_카테고리"].notna()].reset_index(drop=True)
    df["가격_원"] = pd.to_numeric(df["가격_원"], errors="coerce")

    def _to_float_safe(v):
        try:
            return float(v)
        except (ValueError, TypeError):
            return None

    df["두께_mm_num"] = df["두께_mm"].apply(_to_float_safe)
    df["열전도율_num"] = df["열전도율"].apply(_to_float_safe)

    return df


def load_indirect_cost_matrix(xlsx_path: Optional[str] = None) -> dict:
    """08 조달청 간접공사비 매트릭스 -> dict."""
    if xlsx_path is None:
        xlsx_path = os.getenv(
            "INDIRECT_COST_PATH",
            "./data/policy_docs/08_조달청_간접공사비_2026.xlsx",
        )

    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sh = wb["건축제비율(26.1.1.)"]

    # 간접노무비 + 기타경비: 공사규모 × 공사기간 매트릭스
    matrix = {}
    current_scale = None

    for r in range(15, 55):
        scale_cell = sh.cell(r, 2).value
        duration_cell = sh.cell(r, 13).value
        labor = sh.cell(r, 29).value
        expense = sh.cell(r, 38).value

        if scale_cell:
            current_scale = " ".join(str(scale_cell).replace("\n", " ").split())

        if duration_cell and labor is not None and expense is not None:
            duration = " ".join(str(duration_cell).split())
            key = (current_scale, duration)
            matrix[key] = {
                "간접노무": float(labor) / 100,
                "기타경비": float(expense) / 100,
            }

    # 일반관리비 + 이윤: 공사규모 단독
    gm_matrix = {}
    profit_matrix = {}
    current_scale_v = None

    for r in range(13, 50):
        scale = sh.cell(r, 47).value
        gm = sh.cell(r, 57).value
        profit = sh.cell(r, 70).value

        if scale:
            current_scale_v = " ".join(str(scale).replace("\n", " ").split())

        if gm is not None and current_scale_v:
            gm_matrix[current_scale_v] = float(gm) / 100
        if profit is not None and current_scale_v:
            profit_matrix[current_scale_v] = float(profit) / 100

    return {
        "labor_expense": matrix,
        "general_mgmt": gm_matrix,
        "profit": profit_matrix,
    }


# ====================================================================
# Phase B - 자재 조회
# ====================================================================

def lookup_material_price(
    df: pd.DataFrame,
    gr_category: str,
    min_thickness_mm: Optional[float] = None,
    prefer_lowest_price: bool = True,
) -> dict:
    """GR 카테고리에서 조건에 맞는 자재 조회."""
    subset = df[df["GR_카테고리"] == gr_category].copy()

    if subset.empty:
        raise ValueError(f"GR 카테고리 '{gr_category}'에 해당하는 자재 없음")

    if min_thickness_mm is not None:
        subset = subset[subset["두께_mm_num"] >= min_thickness_mm]

    if subset.empty:
        raise ValueError(
            f"'{gr_category}'에 두께 {min_thickness_mm}mm 이상 자재 없음"
        )

    if prefer_lowest_price:
        row = subset.sort_values("가격_원").iloc[0]
    else:
        row = subset.iloc[0]

    return {
        "품명": row["품명"],
        "규격": row["규격_요약"],
        "단가": int(row["가격_원"]) if pd.notna(row["가격_원"]) else 0,
        "단위": row["단위"],
        "두께_mm": row["두께_mm_num"],
        "열전도율": row["열전도율_num"],
    }


# ====================================================================
# Phase C - 직접공사비 BOQ
# ====================================================================

def calculate_direct_cost(
    bim_input: dict,
    price_db: Optional[pd.DataFrame] = None,
) -> dict:
    """BIM 추출 면적 + 자재 단가 -> 직접공사비 BOQ."""
    if price_db is None:
        price_db = load_price_db()

    items = []

    # 1. 외벽 단열
    wall_m2 = bim_input.get("wall_no_insulation_m2", 0)
    if wall_m2 > 0:
        mat = lookup_material_price(
            price_db,
            gr_category=bim_input.get("wall_insulation_material", "GR_단열_PF"),
            min_thickness_mm=bim_input.get("wall_thickness_mm", 130),
        )
        # 자재 단가 × 시공·마감·가설 계수
        unit_price = int(mat["단가"] * CONSTRUCTION_FACTOR["외벽_외단열"])
        items.append({
            "항목": "외벽 단열보강",
            "자재": f"{mat['품명']} {mat['규격']} (자재{mat['단가']:,}원 × 시공계수 {CONSTRUCTION_FACTOR['외벽_외단열']})",
            "수량": wall_m2,
            "단위": "㎡",
            "단가": unit_price,
            "금액": int(wall_m2 * unit_price),
        })

    # 2. 창호
    window_m2 = bim_input.get("window_area_m2", 0)
    if window_m2 > 0:
        mat = lookup_material_price(
            price_db,
            gr_category=bim_input.get("window_material", "GR_창호_복층유리"),
            min_thickness_mm=bim_input.get("window_thickness_mm", 24),
        )
        # 창호는 유리 자재 × 프레임+시공 계수
        unit_price = int(mat["단가"] * CONSTRUCTION_FACTOR["창호"])
        items.append({
            "항목": "창호 교체",
            "자재": f"{mat['품명']} {mat['규격']} (유리{mat['단가']:,}원 × 종합계수 {CONSTRUCTION_FACTOR['창호']})",
            "수량": window_m2,
            "단위": "㎡",
            "단가": unit_price,
            "금액": int(window_m2 * unit_price),
        })

    # 3. 단열문
    door_m2 = bim_input.get("door_area_m2", 0)
    if door_m2 > 0:
        mat = lookup_material_price(price_db, gr_category="GR_문_금속문")
        door_count = bim_input.get("door_count", max(1, int(door_m2 / 2.2)))
        unit_price = int(mat["단가"] * CONSTRUCTION_FACTOR["단열문"])
        items.append({
            "항목": "고기밀성 단열문",
            "자재": f"{mat['품명']} (자재{mat['단가']:,}원 × 시공계수 {CONSTRUCTION_FACTOR['단열문']})",
            "수량": door_count,
            "단위": "개",
            "단가": unit_price,
            "금액": int(door_count * unit_price),
        })

    # 4. 지붕 단열 (선택)
    roof_m2 = bim_input.get("roof_area_m2", 0)
    if roof_m2 > 0:
        mat = lookup_material_price(
            price_db,
            gr_category="GR_단열_XPS",
            min_thickness_mm=180,
        )
        unit_price = int(mat["단가"] * CONSTRUCTION_FACTOR["지붕_단열"])
        items.append({
            "항목": "지붕 단열",
            "자재": f"{mat['품명']} {mat['규격']} × 시공계수 {CONSTRUCTION_FACTOR['지붕_단열']}",
            "수량": roof_m2,
            "단위": "㎡",
            "단가": unit_price,
            "금액": int(roof_m2 * unit_price),
        })

    # 5. 바닥 단열 (선택)
    floor_m2 = bim_input.get("floor_area_m2", 0)
    if floor_m2 > 0:
        mat = lookup_material_price(
            price_db,
            gr_category="GR_단열_경질우레탄",
            min_thickness_mm=50,
        )
        unit_price = int(mat["단가"] * CONSTRUCTION_FACTOR["바닥_단열"])
        items.append({
            "항목": "바닥 단열",
            "자재": f"{mat['품명']} {mat['규격']} × 시공계수 {CONSTRUCTION_FACTOR['바닥_단열']}",
            "수량": floor_m2,
            "단위": "㎡",
            "단가": unit_price,
            "금액": int(floor_m2 * unit_price),
        })

    direct_total = sum(item["금액"] for item in items)

    return {
        "items": items,
        "직접공사비_합계": direct_total,
    }


# ====================================================================
# Phase D - 간접비 적용
# ====================================================================

def _get_scale_key(direct_cost: float) -> str:
    """직접공사비 -> 공사규모 구간 키 (간접노무·기타경비용)."""
    eok = direct_cost / 100_000_000
    if eok < 10:
        return "10억 미만"
    elif eok < 50:
        return "10억 - 50억 미만"
    elif eok < 300:
        return "50억 - 300억 미만"
    elif eok < 1000:
        return "300억 - 1000억 미만"
    return "1000억 이상"


def _get_scale_key_v(direct_cost: float) -> str:
    """일반관리비·이윤 매트릭스용 공사규모 키."""
    eok = direct_cost / 100_000_000
    if eok < 5:
        return "5억 미만"
    elif eok < 30:
        return "5억 - 30억 미만"
    elif eok < 50:
        return "30억 - 50억 미만"
    elif eok < 100:
        return "50억 - 100억 미만"
    elif eok < 300:
        return "100억 - 300억 미만"
    elif eok < 1000:
        return "300억 - 1000억 미만"
    return "1000억 이상"


def _get_duration_key(months: int) -> str:
    if months <= 6:
        return "6개월 이하 (183일)"
    elif months <= 12:
        return "7~12개월 (365일)"
    elif months <= 36:
        return "13~36개월 (1095일)"
    return "36개월 초과 (1096일)"


def apply_indirect_cost(
    direct_cost: float,
    project_duration_months: int = 8,
    matrix: Optional[dict] = None,
) -> dict:
    """
    08 조달청 간접공사비 매트릭스 적용.

    산정식:
        간접노무비 = 직접노무비 × 율 (직노 = 직접비 × 0.3 가정)
        기타경비   = (재+노) × 율 ≈ 직접비 × 율
        일반관리비 = (재+노+경) × 율
        이윤       = (노+경+일) × 율
        공급가액   = 직접비 + 간접노무 + 기타경비 + 일반관리비 + 이윤
        부가세     = 공급가액 × 10%
        Max Cost   = 공급가액 + 부가세
    """
    if matrix is None:
        matrix = load_indirect_cost_matrix()

    scale_lr = _get_scale_key(direct_cost)
    duration = _get_duration_key(project_duration_months)
    scale_gm = _get_scale_key_v(direct_cost)

    le = matrix["labor_expense"].get((scale_lr, duration), {
        "간접노무": 0.152, "기타경비": 0.047,
    })
    gm_rate = matrix["general_mgmt"].get(scale_gm, 0.08)
    profit_rate = matrix["profit"].get(scale_gm, 0.15)

    labor_share = 0.30   # 직노 비중 가정
    direct_labor = direct_cost * labor_share

    indirect_labor = direct_labor * le["간접노무"]
    other_expense = direct_cost * le["기타경비"]

    rne = direct_cost + indirect_labor + other_expense
    general_mgmt = rne * gm_rate

    noe_gm = direct_labor + indirect_labor + other_expense + general_mgmt
    profit = noe_gm * profit_rate

    supply_value = direct_cost + indirect_labor + other_expense + general_mgmt + profit
    vat = supply_value * 0.10
    max_cost = supply_value + vat

    return {
        "직접공사비": int(direct_cost),
        "간접노무비": int(indirect_labor),
        "기타경비": int(other_expense),
        "일반관리비": int(general_mgmt),
        "이윤": int(profit),
        "공급가액": int(supply_value),
        "부가세": int(vat),
        "Max_Cost": int(max_cost),
        "_적용율": {
            "간접노무율": le["간접노무"],
            "기타경비율": le["기타경비"],
            "일반관리율": gm_rate,
            "이윤율": profit_rate,
        },
    }


# ====================================================================
# Phase E - 보조금 · 인센티브
# ====================================================================

def calculate_subsidy(
    max_cost: float,
    total_area_m2: float,
    is_seoul_or_public: bool = True,
    is_signature: bool = False,
) -> dict:
    """GR 국비 보조금 산정."""
    pyeong = total_area_m2 / 3.3
    is_small = total_area_m2 < 300

    if is_seoul_or_public:
        rate = SUBSIDY_RATE["서울_중앙_공공"]
        cap = SUBSIDY_PROJECT_CAP["서울_중앙_공공"]
        if is_signature:
            key = "서울_중앙_공공_시그니처"
        elif is_small:
            key = "서울_중앙_공공_소규모"
        else:
            key = "서울_중앙_공공_일반"
    else:
        rate = SUBSIDY_RATE["그외_지자체"]
        cap = SUBSIDY_PROJECT_CAP["그외_지자체"]
        if is_signature:
            key = "그외_지자체_시그니처"
        elif is_small:
            key = "그외_지자체_소규모"
        else:
            key = "그외_지자체_일반"

    unit_limit = SUBSIDY_LIMIT_PER_PYEONG[key] * 10_000
    area_limit = unit_limit * pyeong
    subsidy_by_rate = max_cost * rate
    subsidy = min(subsidy_by_rate, area_limit, cap)
    self_burden = max_cost - subsidy

    return {
        "보조율": rate,
        "단위면적당_한도_원_per_3.3m2": unit_limit,
        "면적_한도": int(area_limit),
        "보조율_기반_금액": int(subsidy_by_rate),
        "사업당_상한": cap,
        "보조금": int(subsidy),
        "자부담": int(self_burden),
    }


def calculate_far_bonus(
    base_area_m2: float,
    zeb_grade: int,
    land_price_per_pyeong: float,
) -> dict:
    """용적률 완화 인센티브 자산가치 (녹색건축법 §15)."""
    bonus_rate = FAR_BONUS_BY_GRADE.get(zeb_grade, 0.0)
    extra_area_m2 = base_area_m2 * bonus_rate
    extra_pyeong = extra_area_m2 / 3.3
    asset_value = extra_pyeong * land_price_per_pyeong

    return {
        "용적률_보너스율": bonus_rate,
        "추가_연면적_m2": round(extra_area_m2, 2),
        "추가_평수": round(extra_pyeong, 2),
        "평당가": land_price_per_pyeong,
        "자산가치": int(asset_value),
    }


def calculate_acquisition_tax_relief(
    build_cost: float,
    zeb_grade: int,
) -> dict:
    """취득세 감면액 (지방세특례 §47의2)."""
    relief_rate = ZEB_TAX_RELIEF_RATE.get(zeb_grade, 0.0)
    tax_base = build_cost * ACQUISITION_TAX_RATE
    relief = tax_base * relief_rate

    return {
        "신축비": int(build_cost),
        "취득세율": ACQUISITION_TAX_RATE,
        "취득세": int(tax_base),
        "감면율": relief_rate,
        "감면액": int(relief),
    }


# ====================================================================
# Phase E2 - 현금흐름 수익성 (NPV / IRR / B-C / 할인회수)
# ====================================================================
# 정적 회수기간(자부담 ÷ 연절감)은 화폐의 시간가치를 무시합니다.
# 공공시설(어린이집) 경제성 평가는 예비타당성조사(KDI) 방식의
# NPV / B-C / 할인회수기간이 표준입니다.
#
# 기본 가정 (모두 출처 있는 보수적 값, 노란 셀처럼 조정 가능):
#   - 할인율 4.5%      : KDI 사회적 할인율 (2017~, 비SOC 4.5%)
#   - 에너지상승률 2.5%: 전기·도시가스 장기 평균 추세 (보수적)
#   - 분석기간 20년    : 외피 단열·창호 내용연수 기준
# --------------------------------------------------------------------

def _npv_at(rate: float, flows: list) -> float:
    """현금흐름 벡터의 순현재가치 (flows[0] = 0년차)."""
    return sum(cf / ((1 + rate) ** i) for i, cf in enumerate(flows))


def _irr(flows: list, lo: float = -0.9, hi: float = 5.0,
         tol: float = 1e-8, maxit: int = 300) -> Optional[float]:
    """이분법 내부수익률 (의존성 없이 IRR 산출). 부호변화 없으면 None."""
    f_lo = _npv_at(lo, flows)
    f_hi = _npv_at(hi, flows)
    if f_lo * f_hi > 0:
        return None
    for _ in range(maxit):
        mid = (lo + hi) / 2.0
        f = _npv_at(mid, flows)
        if abs(f) < tol or (hi - lo) < tol:
            return mid
        if f_lo * f < 0:
            hi = mid
        else:
            lo = mid
            f_lo = f
    return (lo + hi) / 2.0


def calculate_cashflow_metrics(
    self_burden: float,
    annual_saving: float,
    analysis_years: int = 20,
    energy_escalation: float = 0.025,
    discount_rate: float = 0.045,
    annual_maintenance: float = 0.0,
) -> Optional[dict]:
    """
    GR 투자 현금흐름 기반 수익성 지표.

    - 투자: 0년차에 자부담(Equity) 유출
    - 편익: 매년 에너지 절감액(상승률 반영), 유지보수비 차감
    - 할인: 사회적 할인율로 현재가치 환산
    → NPV / IRR / B-C / 단순·할인 회수기간
    """
    if not self_burden or self_burden <= 0 or not annual_saving or annual_saving <= 0:
        return None

    self_burden = float(self_burden)
    annual_saving = float(annual_saving)

    flows = [-self_burden]
    pv_benefit = nom_benefit = cum_nom = cum_disc = 0.0
    nom_payback = disc_payback = None
    yearly = []

    for t in range(1, analysis_years + 1):
        saving = annual_saving * ((1 + energy_escalation) ** (t - 1)) - annual_maintenance
        flows.append(saving)
        disc = saving / ((1 + discount_rate) ** t)
        pv_benefit += disc
        nom_benefit += saving
        prev_nom, prev_disc = cum_nom, cum_disc
        cum_nom += saving
        cum_disc += disc
        if nom_payback is None and cum_nom >= self_burden:
            nom_payback = (t - 1) + (self_burden - prev_nom) / saving
        if disc_payback is None and cum_disc >= self_burden:
            disc_payback = (t - 1) + (self_burden - prev_disc) / disc
        yearly.append({
            "연차": t,
            "절감액": int(saving),
            "할인편익": int(disc),
            "누적할인편익": int(cum_disc),
        })

    npv = pv_benefit - self_burden
    irr = _irr(flows)

    return {
        "분석기간_년": analysis_years,
        "할인율": discount_rate,
        "에너지상승률": energy_escalation,
        "자부담_원": int(self_burden),
        "명목총절감_원": int(nom_benefit),
        "편익현재가치_원": int(pv_benefit),
        "NPV_원": int(npv),
        "BC_ratio": round(pv_benefit / self_burden, 2),
        "IRR": round(irr, 4) if irr is not None else None,
        "단순회수_년": round(nom_payback, 1) if nom_payback else None,
        "할인회수_년": round(disc_payback, 1) if disc_payback else None,
        "_yearly": yearly,
    }


# ====================================================================
# Phase F - 통합 ROI
# ====================================================================

def calculate_roi(
    bim_input: dict,
    building_info: dict,
) -> dict:
    """통합 ROI 산출 - 모드 2의 메인 진입점."""
    # Phase C: 직접공사비
    boq = calculate_direct_cost(bim_input)

    # Phase D: 간접비 -> Max Cost
    indirect = apply_indirect_cost(
        boq["직접공사비_합계"],
        project_duration_months=building_info.get("project_duration_months", 8),
    )
    max_cost = indirect["Max_Cost"]

    # Phase E: 보조금
    subsidy_info = calculate_subsidy(
        max_cost,
        building_info["total_area_m2"],
        is_seoul_or_public=building_info.get("is_seoul_or_public", True),
        is_signature=building_info.get("is_signature", False),
    )

    # 증축동 신축비
    ext_m2 = building_info.get("extension_area_m2", 0)
    ext_pyeong = ext_m2 / 3.3
    build_cost = ext_pyeong * building_info.get("build_cost_per_pyeong", 9_750_000)

    # 용적률 보너스
    far = calculate_far_bonus(
        ext_m2,
        building_info.get("zeb_target_grade", 5),
        building_info.get("land_price_per_pyeong", 15_000_000),
    )

    # 취득세 감면
    tax = calculate_acquisition_tax_relief(
        build_cost,
        building_info.get("zeb_target_grade", 5),
    )

    # 연간 에너지 절감액
    annual_saving_per_m2 = building_info.get("annual_energy_saving_won_per_m2", 9_900)
    annual_saving = building_info["total_area_m2"] * annual_saving_per_m2

    # 현금흐름 기반 수익성 (NPV / IRR / B-C / 할인회수) — 자부담 기준
    cashflow = calculate_cashflow_metrics(
        self_burden=subsidy_info["자부담"],
        annual_saving=annual_saving,
        analysis_years=building_info.get("analysis_years", 20),
        energy_escalation=building_info.get("energy_escalation", 0.025),
        discount_rate=building_info.get("discount_rate", 0.045),
    )

    # ROI 산출
    total_investment = subsidy_info["자부담"] + build_cost
    immediate_benefit = far["자산가치"] + tax["감면액"]

    # 자산화 ROI: 신축비 대비 즉시 자산가치
    asset_roi_pct = (far["자산가치"] / build_cost * 100) if build_cost > 0 else 0

    # 회수기간 두 가지로 분리 표시
    # 1. GR 단독 회수기간 (자부담 / 연간 절감)
    gr_only_payback = (
        subsidy_info["자부담"] / annual_saving
    ) if annual_saving > 0 else None

    # 2. 통합 회수기간 (총 투자 - 즉시 수혜) / 연간 절감
    net_investment = total_investment - immediate_benefit
    combined_payback = (
        net_investment / annual_saving
    ) if annual_saving > 0 else None

    return {
        "boq": boq,
        "indirect": indirect,
        "max_cost": max_cost,
        "subsidy": subsidy_info,
        "build_cost": int(build_cost),
        "far_bonus": far,
        "tax_relief": tax,
        "annual_saving": int(annual_saving),
        "total_investment": int(total_investment),
        "immediate_benefit": int(immediate_benefit),
        "net_investment": int(net_investment),
        "asset_roi_pct": round(asset_roi_pct, 2),
        "gr_only_payback_years": round(gr_only_payback, 1) if gr_only_payback else None,
        "combined_payback_years": round(combined_payback, 1) if combined_payback else None,
        "cashflow": cashflow,
    }
